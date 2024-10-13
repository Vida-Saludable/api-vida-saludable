import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.db.models import Q
from rest_framework.views import APIView
from users.models.datos_personales_usuario_model import DatosPersonalesUsuario
from habits.models.dormir_model import Dormir

class ClasificacionDormirUsuariosExcelAPIView(APIView):
    def get(self, request, *args, **kwargs):
        # Obtener parámetros de filtro de la URL
        hora_dormir = request.query_params.get('hora', None)
        fecha_inicio = request.query_params.get('fecha_inicio', None)
        fecha_fin = request.query_params.get('fecha_fin', None)

        # Filtrar los registros de Dormir según los criterios recibidos
        filtros = Q()
        if hora_dormir:
            filtros &= Q(hora=hora_dormir)
        if fecha_inicio and fecha_fin:
            filtros &= Q(fecha__range=[fecha_inicio, fecha_fin])

        # Obtener los registros de Dormir con los filtros
        dormir_qs = Dormir.objects.filter(filtros).select_related('usuario')

        # Obtener los IDs de los usuarios de esos registros
        usuario_ids = dormir_qs.values_list('usuario_id', flat=True).distinct()

        # Obtener los datos personales de esos usuarios
        usuarios = DatosPersonalesUsuario.objects.filter(usuario__id__in=usuario_ids)

        # Crear un libro de trabajo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios Dormir"

        # Definir el encabezado de la tabla
        headers = [
            "Nombres Apellidos", "Sexo", "Edad", "Estado Civil",
            "Fecha de Nacimiento", "Teléfono", "Grado de Instrucción",
            "Procedencia", "Religión", "Correo"
        ]

        # Agregar el encabezado en la primera fila con negrita
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)  # Establecer negrita

        # Agregar los datos de los usuarios en las filas siguientes
        for row_num, usuario in enumerate(usuarios, 2):
            ws.cell(row=row_num, column=1, value=usuario.nombres_apellidos)
            ws.cell(row=row_num, column=2, value=usuario.sexo)
            ws.cell(row=row_num, column=3, value=usuario.edad)
            ws.cell(row=row_num, column=4, value=usuario.estado_civil)
            ws.cell(row=row_num, column=5, value=usuario.fecha_nacimiento)
            ws.cell(row=row_num, column=6, value=usuario.telefono)
            ws.cell(row=row_num, column=7, value=usuario.grado_instruccion)
            ws.cell(row=row_num, column=8, value=usuario.procedencia)
            ws.cell(row=row_num, column=9, value=usuario.religion)
            ws.cell(row=row_num, column=10, value=usuario.usuario.correo)

        # Crear una respuesta HTTP para enviar el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=usuarios_dormir.xlsx'
        wb.save(response)

        return response

import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from django.http import HttpResponse
from habits.models.ejercicio_model import Ejercicio
from users.models.datos_personales_usuario_model import DatosPersonalesUsuario

class ExportarEjercicioExcelView(APIView):
    def get(self, request):
        usuario_param = request.query_params.get('usuario', None)
        proyecto_param = request.query_params.get('proyecto', None)

        queryset = Ejercicio.objects.all().order_by('fecha')

        if usuario_param:
            queryset = queryset.filter(usuario__datospersonalesusuario__nombres_apellidos__icontains=usuario_param)

        if proyecto_param:
            queryset = queryset.filter(usuario__usuarioproyecto_set__proyecto__id=proyecto_param).distinct()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ejercicios"

        headers = [
            "ID",
            "Fecha",
            "Usuario",
            "Teléfono",
            "Proyecto",
            "Duración (minutos)",
            "Tipo de Ejercicio",
        ]
        ws.append(headers)

        for ejercicio in queryset:
            usuario = ejercicio.usuario
            datos_personales = DatosPersonalesUsuario.objects.filter(usuario_id=usuario.id).first()
            nombres_apellidos = datos_personales.nombres_apellidos if datos_personales else ""
            telefono = datos_personales.telefono if datos_personales else ""

            proyecto = usuario.usuarioproyecto_set.first()
            nombre_proyecto = proyecto.proyecto.nombre if proyecto else ""

            fila = [
                ejercicio.id,
                ejercicio.fecha.strftime("%Y-%m-%d"),
                nombres_apellidos,
                telefono,
                nombre_proyecto,
                ejercicio.tiempo,
                ejercicio.tipo,
            ]
            ws.append(fila)

        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=ejercicios.xlsx'
        wb.save(response)
        return response

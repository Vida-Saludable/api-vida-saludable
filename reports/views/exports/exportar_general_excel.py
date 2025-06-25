import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime, timedelta
from collections import defaultdict

from rest_framework.views import APIView
from django.db.models import Sum

from users.models.datos_personales_usuario_model import DatosPersonalesUsuario
from users.models.usuario_proyecto_model import UsuarioProyecto
from health.models.datos_fisicos_models import DatosFisicos
from users.models.usuario_models import Usuario

from habits.models.alimentacion_model import Alimentacion
from habits.models.aire_model import Aire
from habits.models.agua_model import Agua
from habits.models.ejercicio_model import Ejercicio
from habits.models.esperanza_model import Esperanza
from habits.models.sol_model import Sol
from habits.models.dormir_model import Dormir
from habits.models.despertar_model import Despertar


class ExportarRegistroHabitosExcelView(APIView):
    def get(self, request):
        fecha_inicial = request.query_params.get('fecha_inicial')
        fecha_final = request.query_params.get('fecha_final')
        proyecto_id = request.query_params.get('proyecto')
        nombre = request.query_params.get('nombre')

        date_filters = {}
        if fecha_inicial:
            date_filters['fecha__gte'] = datetime.strptime(fecha_inicial, "%Y-%m-%d")
        if fecha_final:
            date_filters['fecha__lte'] = datetime.strptime(fecha_final, "%Y-%m-%d")

        usuarios = Usuario.objects.all()
        if nombre:
            usuarios = usuarios.filter(datospersonalesusuario__nombres_apellidos__icontains=nombre)
        if proyecto_id:
            usuarios = usuarios.filter(usuarioproyecto__proyecto_id=proyecto_id)

        usuarios_ids = usuarios.values_list('id', flat=True)
        resultado_agrupado = defaultdict(lambda: defaultdict(dict))

        registros_alimentacion = Alimentacion.objects.filter(usuario_id__in=usuarios_ids, **date_filters).values('fecha', 'usuario_id').annotate(
            desayuno=Sum('desayuno'), almuerzo=Sum('almuerzo'), cena=Sum('cena')
        )
        for r in registros_alimentacion:
            resultado_agrupado[r['usuario_id']][r['fecha']]['alimentacion'] = {
                'desayuno': r['desayuno'] or 0,
                'almuerzo': r['almuerzo'] or 0,
                'cena': r['cena'] or 0,
            }

        registros_agua = Agua.objects.filter(usuario_id__in=usuarios_ids, **date_filters).values('fecha', 'usuario_id').annotate(
            cantidad=Sum('cantidad')
        )
        for r in registros_agua:
            resultado_agrupado[r['usuario_id']][r['fecha']]['agua'] = {'cantidad': r['cantidad'] or 0}

        registros_aire = Aire.objects.filter(usuario_id__in=usuarios_ids, **date_filters).values('fecha', 'usuario_id').annotate(
            tiempo=Sum('tiempo')
        )
        for r in registros_aire:
            resultado_agrupado[r['usuario_id']][r['fecha']]['aire'] = {'tiempo': r['tiempo'] or 0}

        registros_ejercicio = Ejercicio.objects.filter(usuario_id__in=usuarios_ids, **date_filters).values('fecha', 'tipo', 'usuario_id').annotate(
            tiempo=Sum('tiempo')
        )
        for r in registros_ejercicio:
            resultado_agrupado[r['usuario_id']][r['fecha']].setdefault('ejercicio', []).append({
                'tipo': r['tipo'], 'tiempo': r['tiempo']
            })

        registros_esperanza = Esperanza.objects.filter(usuario_id__in=usuarios_ids, **date_filters).values('fecha', 'tipo_practica', 'usuario_id')
        for r in registros_esperanza:
            resultado_agrupado[r['usuario_id']][r['fecha']].setdefault('esperanza', []).append({
                'tipo_practica': r['tipo_practica']
            })

        registros_sol = Sol.objects.filter(usuario_id__in=usuarios_ids, **date_filters).values('fecha', 'usuario_id').annotate(
            tiempo=Sum('tiempo')
        )
        for r in registros_sol:
            resultado_agrupado[r['usuario_id']][r['fecha']]['sol'] = {'tiempo': r['tiempo'] or 0}

        dormir = Dormir.objects.filter(usuario_id__in=usuarios_ids, **date_filters)
        despertar = Despertar.objects.filter(usuario_id__in=usuarios_ids, **date_filters)

        despertar_map = defaultdict(list)
        for d in despertar:
            despertar_map[(d.usuario_id, d.fecha)].append(d.hora)

        for d in dormir:
            uid = d.usuario_id
            f = d.fecha
            h_dormir = datetime.combine(f, d.hora)

            posibles_desp = []
            for offset in [0, 1]:
                f_desp = f + timedelta(days=offset)
                for h in despertar_map.get((uid, f_desp), []):
                    h_desp = datetime.combine(f_desp, h)
                    if h_desp > h_dormir:
                        posibles_desp.append(h_desp)

            if posibles_desp:
                h_despertar = min(posibles_desp)
                total_min = int((h_despertar - h_dormir).total_seconds() / 60)
                horas = total_min // 60
                minutos = total_min % 60
            else:
                horas = 0
                minutos = 0

            resultado_agrupado[uid][f]['descanso'] = {
                'total_horas': horas,
                'total_minutos': minutos
            }

        datos_personales = DatosPersonalesUsuario.objects.filter(usuario_id__in=usuarios_ids)
        datos_dict = {dp.usuario_id: dp for dp in datos_personales}

        proyectos = UsuarioProyecto.objects.select_related('proyecto').filter(usuario_id__in=usuarios_ids)
        proyectos_dict = defaultdict(list)
        for p in proyectos:
            proyectos_dict[p.usuario_id].append(p.proyecto.nombre)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros de Hábitos"

        headers = [
            "Fecha", "Nombre", "Teléfono", "Proyectos", "Peso", 
            "Desayuno", "Almuerzo", "Cena", "Agua (ml)", "Aire (min)", 
            "Ejercicio", "Esperanza", "Sol (min)",
            "Horas Dormidas", "Minutos Dormidos"
        ]
        ws.append(headers)

        for usuario_id, fechas in resultado_agrupado.items():
            for fecha, data in fechas.items():
                dp = datos_dict.get(usuario_id)
                nombre = dp.nombres_apellidos if dp else ""
                telefono = dp.telefono if dp else ""
                peso = DatosFisicos.objects.filter(usuario_id=usuario_id).order_by('-fecha').first()
                peso_valor = peso.peso if peso else ""

                proyectos_str = ", ".join(proyectos_dict[usuario_id])
                ejercicios = ", ".join(f"{e['tipo']} ({e['tiempo']} min)" for e in data.get('ejercicio', []))
                esperanzas = ", ".join(e['tipo_practica'] for e in data.get('esperanza', []))

                descanso = data.get('descanso', {'total_horas': 0, 'total_minutos': 0})

                fila = [
                    fecha.strftime('%Y-%m-%d'),
                    nombre,
                    telefono,
                    proyectos_str,
                    peso_valor,
                    data.get('alimentacion', {}).get('desayuno', 0),
                    data.get('alimentacion', {}).get('almuerzo', 0),
                    data.get('alimentacion', {}).get('cena', 0),
                    data.get('agua', {}).get('cantidad', 0),
                    data.get('aire', {}).get('tiempo', 0),
                    ejercicios,
                    esperanzas,
                    data.get('sol', {}).get('tiempo', 0),
                    descanso['total_horas'],
                    descanso['total_minutos'],
                ]
                ws.append(fila)

        for i, col in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=registro_habitos.xlsx'
        wb.save(response)
        return response

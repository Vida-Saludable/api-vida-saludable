import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from rest_framework.views import APIView
from users.models.datos_personales_usuario_model import DatosPersonalesUsuario
from habits.models.sol_model import Sol

class ClasificacionSolUsuariosExcelAPIView(APIView):
    def get(self, request, *args, **kwargs):
        tiempo = request.query_params.get('tiempo', None)
        fecha_inicio = request.query_params.get('fecha_inicio', None)
        fecha_fin = request.query_params.get('fecha_fin', None)
        filtros = {}
        if tiempo:
            filtros['tiempo'] = tiempo
        if fecha_inicio and fecha_fin:
            filtros['fecha__range'] = [fecha_inicio, fecha_fin]
        sol_qs = Sol.objects.filter(**filtros).select_related('usuario')
        usuario_ids = sol_qs.values_list('usuario_id', flat=True).distinct()
        usuarios = DatosPersonalesUsuario.objects.filter(usuario__id__in=usuario_ids)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios Sol"
        headers = [
            "Nombres Apellidos", "Sexo", "Edad", "Estado Civil",
            "Fecha de Nacimiento", "Teléfono", "Ocupación",
            "Procedencia", "Religión", "Correo"
        ]
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)  

        for row_num, usuario in enumerate(usuarios, 2):
            ws.cell(row=row_num, column=1, value=usuario.nombres_apellidos)
            ws.cell(row=row_num, column=2, value=usuario.sexo)
            ws.cell(row=row_num, column=3, value=usuario.edad)
            ws.cell(row=row_num, column=4, value=usuario.estado_civil)
            ws.cell(row=row_num, column=5, value=usuario.fecha_nacimiento)
            ws.cell(row=row_num, column=6, value=usuario.telefono)
            ws.cell(row=row_num, column=7, value=usuario.ocupacion)
            ws.cell(row=row_num, column=8, value=usuario.procedencia)
            ws.cell(row=row_num, column=9, value=usuario.religion)
            ws.cell(row=row_num, column=10, value=usuario.usuario.correo)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=usuarios_sol.xlsx'
        wb.save(response)

        return response

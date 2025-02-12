import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.db.models import Q
from rest_framework.views import APIView
from users.models.datos_personales_usuario_model import DatosPersonalesUsuario
from habits.models.alimentacion_model import Alimentacion

class ClasificacionAlimentacionUsuariosExcelAPIView(APIView):
    def get(self, request, *args, **kwargs):
        # Obtener parámetros de filtro de la URL
        desayuno_hora = request.query_params.get('desayuno_hora', None)
        almuerzo_hora = request.query_params.get('almuerzo_hora', None)
        cena_hora = request.query_params.get('cena_hora', None)
        desayuno = request.query_params.get('desayuno', None)
        almuerzo = request.query_params.get('almuerzo', None)
        cena = request.query_params.get('cena', None)
        desayuno_saludable = request.query_params.get('desayuno_saludable', None)
        almuerzo_saludable = request.query_params.get('almuerzo_saludable', None)
        cena_saludable = request.query_params.get('cena_saludable', None)
        fecha_inicio = request.query_params.get('fecha_inicio', None)
        fecha_fin = request.query_params.get('fecha_fin', None)

        # Filtrar los registros de Alimentación con los criterios recibidos
        filtros = Q()
        if desayuno_hora:
            filtros &= Q(desayuno_hora=desayuno_hora)
        if almuerzo_hora:
            filtros &= Q(almuerzo_hora=almuerzo_hora)
        if cena_hora:
            filtros &= Q(cena_hora=cena_hora)
        if desayuno:
            filtros &= Q(desayuno=desayuno)
        if almuerzo:
            filtros &= Q(almuerzo=almuerzo)
        if cena:
            filtros &= Q(cena=cena)
        if desayuno_saludable:
            filtros &= Q(desayuno_saludable=desayuno_saludable)
        if almuerzo_saludable:
            filtros &= Q(almuerzo_saludable=almuerzo_saludable)
        if cena_saludable:
            filtros &= Q(cena_saludable=cena_saludable)
        if fecha_inicio and fecha_fin:
            filtros &= Q(fecha__range=[fecha_inicio, fecha_fin])

        # Obtener los registros de alimentación con los filtros (si no hay filtros, trae todos)
        alimentacion_qs = Alimentacion.objects.filter(filtros).select_related('usuario')

        # Obtener los IDs de los usuarios de esos registros
        usuario_ids = alimentacion_qs.values_list('usuario_id', flat=True).distinct()

        # Obtener los datos personales de esos usuarios
        usuarios = DatosPersonalesUsuario.objects.filter(usuario__id__in=usuario_ids)

        # Crear un libro de trabajo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios Alimentación"

        # Definir el encabezado de la tabla
        headers = [
            "Nombres Apellidos", "Sexo", "Edad", "Estado Civil",
            "Fecha de Nacimiento", "Teléfono", "Ocupación",
            "Procedencia", "Religión", "Correo"
        ]

        # Agregar el encabezado en la primera fila con negrita
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)  # Establecer negrita

        # Agregar los datos de los usuarios en las filas siguientes
        for row_num, usuario in enumerate(usuarios, 2):
            ws.cell(row=row_num, column=1, value=usuario.nombres_apellidos)
            ws.cell(row=row_num, column=2, value=usuario.sexo)
            ws.cell(row=row_num, column=3, value=usuario.edad)
            ws.cell(row=row_num, column=4, value=usuario.estado_civil)
            ws.cell(row=row_num, column=5, value=usuario.fecha_nacimiento)
            ws.cell(row=row_num, column=6, value=usuario.telefono)
            ws.cell(row=row_num, column=7, value=usuario.ocupacion)
            ws.cell(row=row_num, column=8, value=usuario.procedencia)
            ws.cell(row=row_num, column=9, value=usuario.religion)
            ws.cell(row=row_num, column=10, value=usuario.usuario.correo)

        # Crear una respuesta HTTP para enviar el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=usuarios_alimentacion.xlsx'
        wb.save(response)

        return response
